#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import csv
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


# Manual override (masalan lab-pan-A uchun)
OVERRIDE_TARGET_CODES = {
    "lab-pan-234": "24356-8",
}


def escape_fsh(value: str) -> str:
    value = "" if value is None else str(value).strip()
    return value.replace("\\", "\\\\").replace('"', '\\"')


def read_excel(path: Path, sheet_name: str | None = None):
    if load_workbook is None:
        raise RuntimeError("pip install openpyxl")

    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    rows = []
    for r in range(2, ws.max_row + 1):
        rows.append({
            "code": str(ws.cell(r, 1).value or "").strip(),
            "uz": str(ws.cell(r, 3).value or "").strip(),
            "ru": str(ws.cell(r, 4).value or "").strip(),
            "loinc": str(ws.cell(r, 6).value or "").strip(),
        })
    return rows


def read_csv(path: Path):
    rows = []
    with path.open("r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        data = list(reader)

    for row in data[1:]:
        rows.append({
            "code": str(row[0] if len(row) > 0 else "").strip(),
            "uz": str(row[2] if len(row) > 2 else "").strip(),
            "ru": str(row[3] if len(row) > 3 else "").strip(),
            "loinc": str(row[5] if len(row) > 5 else "").strip(),
        })
    return rows


def read_source(path: Path, sheet_name=None):
    if path.suffix.lower() == ".csv":
        return read_csv(path)
    return read_excel(path, sheet_name)


def build_block(code, uz, ru, loinc):
    return (
        f'* group.element[+].code = #{escape_fsh(code)}\n'
        f'* group.element[=].display = "{escape_fsh(uz)}"\n'
        f'* group.element[=].target[+].code = #{escape_fsh(loinc)}\n'
        f'* group.element[=].target[=].display = "{escape_fsh(ru)}"\n'
        f'* group.element[=].target[=].relationship = #equivalent\n'
    )


def generate(rows):
    result = []

    for r in rows:
        code = r["code"]
        uz = r["uz"]
        ru = r["ru"]
        loinc = r["loinc"] or OVERRIDE_TARGET_CODES.get(code, "")

        if not code or not uz or not ru or not loinc:
            continue

        result.append(build_block(code, uz, ru, loinc))

    return "\n".join(result)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("input_file")
    parser.add_argument("-o", "--output", default="conceptmap.fsh")
    parser.add_argument("--sheet", default=None)
    args = parser.parse_args()

    rows = read_source(Path(args.input_file), args.sheet)
    fsh = generate(rows)

    Path(args.output).write_text(fsh, encoding="utf-8")
    print("DONE")


if __name__ == "__main__":
    main()